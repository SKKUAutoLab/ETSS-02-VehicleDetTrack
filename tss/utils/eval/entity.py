import os
import sys
import glob

from typing import Optional

import cv2
import numpy as np

import openpyxl as pyxl
from shutil import copyfile


class CountUnit(object):

	def __init__(
			self,
			gen_time        : float = 0.1,
			video_id        : int   = 1,
			frame_id        : int   = 1,
			movement_id     : int   = 1,
			vehicle_class_id: int   = 0,
			**kwargs
	):
		super().__init__(**kwargs)
		self.gen_time = gen_time
		self.video_id = video_id
		self.frame_id = frame_id
		self.movement_id = movement_id
		self.vehicle_class_id = vehicle_class_id

	def __gt__(self, other):

		# TODO: First check frame_id
		if self.frame_id > other.frame_id:
			return True

		# TODO: Second check movement_id
		if self.frame_id == other.frame_id and self.movement_id > other.movement_id:
			return True

		return False

	def __eq__(self, other):
		if self.frame_id == other.frame_id and self.movement_id == other.movement_id:
			return True
		return False

	def __lt__(self, other):
		# TODO: First check frame_id
		if self.frame_id < other.frame_id:
			return True

		# TODO: Second check movement_id
		if self.frame_id == other.frame_id and self.movement_id < other.movement_id:
			return True

		return False


class ListCountUnit(object):

	# MARK: Magic Functions

	def __init__(
			self,
			vehicle_class_id: int = 0,
			**kwargs
	):
		super().__init__(**kwargs)
		self.vehicle_class_id = vehicle_class_id
		self.list_count       = []
		self.count_unit_idx   = 0

	def __len__(self):
		return len(self.list_count)

	def __iter__(self):
		self.count_unit_idx = -1
		return self

	def __next__(self):
		if self.count_unit_idx >= len(self.list_count) - 1:
			raise StopIteration
		else:
			self.count_unit_idx += 1
			return self.list_count[self.count_unit_idx]

	def __del__(self):
		del self.list_count

	# MARK: ADJUSTING

	def add_count_unit_(self, count_unit_: CountUnit):
		self.list_count.append(count_unit_)
		self.list_count.sort()

	def remove_count_unit_(self, count_unit_: CountUnit):
		for count_unit in self.list_count:
			if count_unit == count_unit_:
				self.list_count.remove(count_unit)
				break


class FileCount(object):

	def __init__(
			self,
			path_file   : str            = None,
			path_images : Optional[str]  = None,
			video_info  : dict           = None,
			**kwargs
	):
		super().__init__(**kwargs)
		self.file_name            = os.path.basename(path_file)
		self.path_file            = path_file
		self.video_info           = video_info
		self.list_of_list_count   = []
		self.frame_id_current     = 0  # Index of current frame we focus

		if path_images is not None:
			self.list_imgs = sorted(glob.glob(os.path.join(path_images, "*.jpeg")))

		if path_file.endswith('csv'):
			self.parse_csv(path_file)
		elif path_file.endswith('txt'):
			self.parse_txt(path_file)
		elif path_file.endswith('xlsx'):
			self.parse_xlsx(path_file)

	# MARK: PARSING

	def parse_xlsx(self,xlsx_name: str):
		work_book = pyxl.load_workbook(xlsx_name)
		sheet_names = work_book.sheetnames
		work_sheet = work_book[sheet_names[0]]
		for i in range(2, work_sheet.max_row + 1):
			cvalue = work_sheet.cell(row=i, column=2).value
			if cvalue is not None:
				frame_id = int(cvalue)
			else:
				import pdb;
				pdb.set_trace()
			cvalue = work_sheet.cell(row=i, column=3).value
			if cvalue is not None:
				movement_id = int(cvalue)
			else:
				import pdb;
				pdb.set_trace()
			cvalue = work_sheet.cell(row=i, column=4).value
			if cvalue is not None:
				if cvalue == "car" or cvalue == "car ":
					type = 0
				elif cvalue == "truck" or cvalue == "truch":
					type = 1
				else:
					import pdb;
					pdb.set_trace()

			self.add_count_unit_(CountUnit(
				gen_time         = 0.0,
				video_id         = self.video_info['id'],
				frame_id         = frame_id,
				movement_id      = movement_id,
				vehicle_class_id = type
			))

	def parse_csv(self,	csv_name: str):
		# print(csv_name)
		file_open = open(csv_name)
		lines = file_open.readlines()
		for idx, line in enumerate(lines):

			words = line.rstrip("\n").split(",")

			# NOTE: remove head of each file by checking words[0] is digit or not
			if isinstance(words[0], str):
				if not words[0].lstrip('-').replace('.', '').isdigit():
					continue

			self.add_count_unit_(CountUnit(
				gen_time         = float(words[0]),
				video_id         = int(words[1]),
				frame_id         = int(words[2]),
				movement_id      = int(words[3]),
				vehicle_class_id = int(words[4]) - 1
			))

	def parse_txt(self,	txt_name: str):
		file_open = open(txt_name)

		lines = file_open.readlines()

		for line in lines:
			words = line.rstrip("\n").split(" ")

			# DEBUG:
			# Reduce effectiveness
			# if idx % 4 == 0:
			# continue

			# NOTE: remove head of each file by checking words[0] is digit or not
			if isinstance(words[0], str):
				if not words[0].lstrip('-').replace('.', '').isdigit():
					continue

			self.add_count_unit_(CountUnit(
				gen_time         = float(words[0]),
				video_id         = int(words[1]),
				frame_id         = int(words[2]),
				movement_id      = int(words[3]),
				vehicle_class_id = int(words[4]) - 1
			))

	# MARK: ADJUSTING

	# NOTE: Return the exactly count_unit or less than unit:
	def find_count_unit(self, count_unit_: CountUnit):
		if count_unit_ is None:
			return None

		for list_count in self.list_of_list_count:
			if count_unit_.vehicle_class_id == list_count.vehicle_class_id:
				count_unit_temp = list_count[0]

				for count_unit in list_count:
					if count_unit == count_unit_:
						return count_unit

					if count_unit < count_unit_:
						count_unit_temp = count_unit

					if count_unit > count_unit_:
						break

				return count_unit_temp
		# NOTE: mean none of object with same class_id was found
		return None

	def add_count_unit_(self, count_unit_: CountUnit):
		if len(self.list_of_list_count) > 0:
			for list_count in self.list_of_list_count:
				if list_count.vehicle_class_id == count_unit_.vehicle_class_id:
					list_count.add_count_unit_(count_unit_)
					return

		# NOTE: if len==0 or class_id not in list_count
		self.list_of_list_count.append(ListCountUnit(count_unit_.vehicle_class_id))
		self.list_of_list_count[-1].add_count_unit_(count_unit_)

	def remove_count_unit_(self, count_unit_: CountUnit):
		count_unit_found = self.find_count_unit(count_unit_)

		if count_unit_found is None:
			return

		if len(self.list_of_list_count) > 0:
			for list_count in self.list_of_list_count:
				if list_count.vehicle_class_id == count_unit_found.vehicle_class_id:
					list_count.remove_count_unit_(count_unit_found)

	# MARK: IMAGE SELECTION

	def next_image(self):
		self.frame_id_current += 1
		if self.frame_id_current >= len(self.list_imgs):
			self.frame_id_current = 0

		return self.get_image_current()

	def previous_image(self):
		self.frame_id_current -= 1
		if self.frame_id_current < 0:
			self.frame_id_current = len(self.list_imgs) - 1

		return self.get_image_current()


	# MARK: OUTPUT

	def get_image_current(self) -> np.ndarray:
		"""

		Returns:
			current_image (np.ndarray):
				current processing image
		"""
		# NOTE: frame_id_current == name of images - 1
		print(f"SHOW {self.file_name} :: {self.frame_id_current + 1} :: {self.list_imgs[self.frame_id_current]}")
		return cv2.imread(self.list_imgs[self.frame_id_current])

	def get_array_count(self, total_frame_count : int = -1):
		# TODO: check number of frame to get
		if 0 > total_frame_count or total_frame_count > int(self.video_info['frame_num']):
			total_frame_count = int(self.video_info['frame_num'])

		arr_count = np.zeros((total_frame_count, int(self.video_info['movement_num']), 2), dtype=np.int32)

		# TODO: add count value
		for list_count in self.list_of_list_count:
			for count_unit in list_count:
				if count_unit.frame_id <= total_frame_count:
					arr_count[count_unit.frame_id - 1, count_unit.movement_id - 1, count_unit.vehicle_class_id] += 1

		return arr_count

	def write_file_txt(self):

		file_new = f"{os.path.splitext(self.path_file)[0]}.txt"

		with open(file_new, "w") as file_write:
			arr_count = self.get_array_count()

			for frame_id in range(arr_count.shape[0]):
				for movement_id in range(arr_count.shape[1]):
					for class_id in range(arr_count.shape[2]):
						for num_vehicle in range(arr_count[frame_id, movement_id, class_id]):
							# 〈gen_time〉 〈video_id〉 〈frame_id〉 〈movement_id〉 〈vehicle_class_id〉
							file_write.write(f"0.1 {self.video_info['id']} {frame_id + 1} {movement_id + 1} {class_id + 1}\n")


if __name__ == "__main__":
	pass
